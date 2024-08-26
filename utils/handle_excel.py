import openpyxl

# 读取excel
class ReadExcel():
    def __init__(self, file_path):
        self.workbook = openpyxl.load_workbook(file_path)
        self.worksheet = self.workbook.active

    def get_data(self):
        data = []
        for row in self.worksheet.iter_rows():
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            data.append(row_value)
        return data 
    

class WriteExcel():
    def __init__(self, file_path, data):
        self.file_path = file_path
        self.data = data
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

    def write_to_excel(self):
        for row in self.data:
            self.worksheet.append(row)
        self.workbook.save(self.file_path)


    
    
if __name__ == '__main__':
    import datetime
    # read_excel_obj =  ReadExcel('/Users/andy/Downloads/1年1班学生.xlsx')
    # result = read_excel_obj.get_data()
    # print(result)
    data = ['1年1班', '朱玉', 'G655308201110037589', 'M', datetime.datetime(2011, 10, 3, 0, 0), '18626593089', '湖南省淑英县牧野郭路W座 969796'], ['1年1班', '张桂香', 'G586476201508266228', 'F', datetime.datetime(2015, 8, 26, 0, 0), '15239175647', '河南省刚县合川严路J座 180114'], ['1年1班', '王淑华', 'G629116201305062115', 'F', datetime.datetime(2013, 5, 6, 0, 0), '13483091041', '广西壮族自治区东莞县梁平赵街t座 127684']
    write_excel_obj = WriteExcel('/Users/andy/Downloads/1年1班学生测试.xlsx', data)
    write_excel_obj.write_to_excel()